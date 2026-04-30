# booking/management/commands/import_excel.py
from __future__ import annotations

from pathlib import Path
from datetime import datetime, date, time, timedelta

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

# Modelos
from booking.models import City, Company, Bus, Route, Trip

# ---- helpers comunes ---------------------------------------------------------
def _slugify(s: str) -> str:
    import unicodedata, re
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s

def _norm_key(s: str) -> str:
    import unicodedata
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii","ignore").decode("ascii").strip().lower()

def _money_to_decimal(val):
    from decimal import Decimal
    if val in (None, ""):
        return Decimal("0")
    s = str(val).replace("$", "").replace(" ", "")
    # normaliza separadores (2.500 -> 2500 ; 2,500 -> 2.500)
    s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")

def _to_int(val, default=0):
    try:
        return int(str(val).strip())
    except Exception:
        return default

def _parse_hhmm(v) -> time | None:
    if hasattr(v, "hour") and hasattr(v, "minute"):
        # ya es datetime.time
        return v
    import re
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", str(v))
    if not m:
        return None
    return time(hour=int(m.group(1)), minute=int(m.group(2)))

# ---- importadores por hojas "técnicas" (opcionales) --------------------------
def import_cities(ws, stdout):
    headers = { _norm_key(c): i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) }
    name_idx = headers.get("name") or headers.get("nombre")
    slug_idx = headers.get("slug")
    if name_idx is None:
        stdout.write("  * Hoja 'Ciudades' sin columna 'name/nombre' (se omite).\n")
        return
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[name_idx] or "").strip()
        if not name:
            continue
        slug = (row[slug_idx] or "") if slug_idx is not None else _slugify(name)
        City.objects.get_or_create(name=name, defaults={"slug": slug})
        created += 1
    stdout.write(f"  - Ciudades procesadas: {created}\n")

def import_routes(ws, stdout):
    headers = { _norm_key(c): i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) }
    i_o = headers.get("origin_slug") or headers.get("origen")
    i_d = headers.get("destination_slug") or headers.get("destino")
    i_dur = headers.get("duration_minutes") or headers.get("tiempo")
    i_price = headers.get("base_price") or headers.get("valor")
    if i_o is None or i_d is None:
        stdout.write("  * Hoja 'Rutas' sin origin/destination (se omite).\n")
        return

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        o = (row[i_o] or "").strip()
        d = (row[i_d] or "").strip()
        if not o or not d:
            continue
        origin = City.objects.filter(slug=_slugify(o)).first() or City.objects.filter(name=o).first()
        dest   = City.objects.filter(slug=_slugify(d)).first() or City.objects.filter(name=d).first()
        if not origin or not dest:
            continue
        dur = _to_int(row[i_dur]) if i_dur is not None else 0
        price = _money_to_decimal(row[i_price]) if i_price is not None else 0
        Route.objects.get_or_create(
            origin=origin, destination=dest,
            defaults={"duration_minutes": dur, "base_price": price}
        )
        created += 1
    stdout.write(f"  - Rutas procesadas: {created}\n")

def import_trips(ws, stdout):
    headers = { _norm_key(c): i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) }
    i_route_o = headers.get("origin_slug") or headers.get("origen")
    i_route_d = headers.get("destination_slug") or headers.get("destino")
    i_date = headers.get("date") or headers.get("fecha")
    i_dep  = headers.get("depart_time") or headers.get("salida")
    i_bus  = headers.get("bus_plate") or headers.get("bus")
    if i_route_o is None or i_route_d is None or i_date is None or i_dep is None:
        stdout.write("  * Hoja 'Viajes' sin columnas suficientes (se omite).\n")
        return

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        o = (row[i_route_o] or "").strip()
        d = (row[i_route_d] or "").strip()
        if not o or not d:
            continue

        origin = City.objects.filter(slug=_slugify(o)).first() or City.objects.filter(name=o).first()
        dest   = City.objects.filter(slug=_slugify(d)).first() or City.objects.filter(name=d).first()
        if not origin or not dest:
            continue

        route = Route.objects.filter(origin=origin, destination=dest).first()
        if not route:
            continue

        # fecha + hora
        dep_date = row[i_date]
        if isinstance(dep_date, str):
            dep_date = datetime.strptime(dep_date, "%Y-%m-%d").date()
        dep_time = _parse_hhmm(row[i_dep])
        if not (dep_date and dep_time):
            continue

        departure_dt = datetime.combine(dep_date, dep_time)
        arrival_dt = departure_dt + timedelta(minutes=(route.duration_minutes or 0))

        bus = None
        if i_bus is not None and row[i_bus]:
            bus = Bus.objects.filter(plate=str(row[i_bus]).strip()).first()

        Trip.objects.get_or_create(
            route=route,
            departure=departure_dt,
            defaults={"arrival": arrival_dt, "bus": bus, "seats_total": getattr(bus, "seats_total", 0) or 0},
        )
        created += 1

    stdout.write(f"  - Viajes creados/ya existentes: {created}\n")

# ---- importador "legacy" (tu hoja con Origen/Destino/Horas/...) ---------------
def import_legacy_routes_and_trips(ws, *, service_date: date | None, default_company_name: str | None,
                                   default_bus_plate: str | None, stdout):
    """
    Columnas esperadas (flexibles por nombre):
      Origen | Destino | Horas | Tiempo | Valor | Terminales | Tipo de Bus
    Crea/actualiza Route (duration/base_price) y, si service_date, crea Trips ese día.
    """
    try:
        from booking.models import Terminal  # si no tienes Terminal, ignora el bloque que lo usa
    except Exception:
        Terminal = None  # noqa

    headers = { _norm_key(c): i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) }
    must = ["origen", "destino"]
    for k in must:
        if k not in headers:
            stdout.write(f"  * Falta columna '{k}' en hoja '{ws.title}' (se omite).\n")
            return

    i_origen   = headers.get("origen")
    i_destino  = headers.get("destino")
    i_horas    = headers.get("horas")
    i_tiempo   = headers.get("tiempo")
    i_valor    = headers.get("valor")
    i_term     = headers.get("terminales")
    # i_tipo_bus = headers.get("tipo de bus")  # reservado para futuras reglas

    # company/bus opcionales
    company = None
    if default_company_name:
        company, _ = Company.objects.get_or_create(name=default_company_name)

    bus = None
    if default_bus_plate:
        bus = Bus.objects.filter(plate=default_bus_plate).first()

    n_routes, n_trips = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        origen  = (row[i_origen]  or "").strip()
        destino = (row[i_destino] or "").strip()
        if not origen or not destino:
            continue

        origin_city, _ = City.objects.get_or_create(name=origen, defaults={"slug": _slugify(origen)})
        dest_city, _   = City.objects.get_or_create(name=destino, defaults={"slug": _slugify(destino)})

        duration_minutes = _to_int(row[i_tiempo]) if i_tiempo is not None else 0
        base_price = _money_to_decimal(row[i_valor]) if i_valor is not None else 0

        route, created = Route.objects.get_or_create(
            origin=origin_city, destination=dest_city,
            defaults={"duration_minutes": duration_minutes, "base_price": base_price}
        )
        if not created:
            changed = False
            if duration_minutes and route.duration_minutes != duration_minutes:
                route.duration_minutes = duration_minutes; changed = True
            if base_price and getattr(route, "base_price", None) != base_price:
                route.base_price = base_price; changed = True
            if changed:
                route.save()
        else:
            n_routes += 1

        # Terminal (si existe el modelo)
        if Terminal and i_term is not None and row[i_term]:
            term_name = str(row[i_term]).strip()
            if term_name:
                Terminal.objects.get_or_create(city=origin_city, name=term_name)

        # Trips
        if service_date and i_horas is not None and row[i_horas]:
            hhmm = _parse_hhmm(row[i_horas])
            if not hhmm:
                continue
            dep = datetime.combine(service_date, hhmm)
            arr = dep + timedelta(minutes=(route.duration_minutes or 0))
            Trip.objects.get_or_create(
                route=route, departure=dep,
                defaults={"arrival": arr, "bus": bus, "seats_total": getattr(bus, "seats_total", 0) or 0},
            )
            n_trips += 1

    stdout.write(f"  - (Legacy) Rutas nuevas: {n_routes}; Viajes creados: {n_trips}\n")

# ---- Management Command -------------------------------------------------------
class Command(BaseCommand):
    help = "Importa datos desde un Excel. Soporta hojas técnicas (Ciudades, Rutas, Viajes) y hoja legacy con Origen/Destino/Horas/…"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Ruta al archivo .xlsx")
        # nombres de hojas "técnicas"
        parser.add_argument("--cities-sheet", default="Ciudades")
        parser.add_argument("--routes-sheet", default="Rutas")
        parser.add_argument("--trips-sheet",  default="Viajes")
        # modo legacy (tu hoja “Rutas” con Origen/Destino/Horas/…)
        parser.add_argument("--legacy-sheet", default="", help="Nombre de hoja con columnas Origen/Destino/Horas/Tiempo/Valor/Terminales.")
        parser.add_argument("--service-date", default="", help="YYYY-MM-DD para crear viajes del modo legacy.")
        parser.add_argument("--company", default="", help="Empresa por defecto (modo legacy, opcional).")
        parser.add_argument("--bus", default="", help="Patente de bus por defecto (modo legacy, opcional).")

    @transaction.atomic
    def handle(self, *args, **opts):
        xlsx_path = Path(opts["xlsx_path"])
        if not xlsx_path.exists():
            raise CommandError(f"No se encontró el archivo: {xlsx_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("Instala openpyxl: pip install openpyxl")

        wb = load_workbook(xlsx_path, data_only=True)

        # 1) Hojas técnicas (si existen)
        cs = opts.get("cities_sheet") or "Ciudades"
        if cs in wb.sheetnames:
            self.stdout.write(self.style.NOTICE(f"Procesando hoja '{cs}'…"))
            import_cities(wb[cs], self.stdout)
        else:
            self.stdout.write(self.style.WARNING(f"Hoja '{cs}' no encontrada (se omite)."))

        rs = opts.get("routes_sheet") or "Rutas"
        if rs in wb.sheetnames:
            self.stdout.write(self.style.NOTICE(f"Procesando hoja '{rs}'…"))
            import_routes(wb[rs], self.stdout)
        else:
            self.stdout.write(self.style.WARNING(f"Hoja '{rs}' no encontrada (se omite)."))

        ts = opts.get("trips_sheet") or "Viajes"
        if ts in wb.sheetnames:
            self.stdout.write(self.style.NOTICE(f"Procesando hoja '{ts}'…"))
            import_trips(wb[ts], self.stdout)
        else:
            self.stdout.write(self.style.WARNING(f"Hoja '{ts}' no encontrada (se omite)."))

        # 2) Modo legacy (tu formato)
        legacy_title = opts.get("legacy_sheet") or ""
        if legacy_title:
            if legacy_title in wb.sheetnames:
                srv_date = None
                if opts.get("service_date"):
                    srv_date = datetime.strptime(opts["service_date"], "%Y-%m-%d").date()
                self.stdout.write(self.style.NOTICE(f"Procesando hoja legacy '{legacy_title}'…"))
                import_legacy_routes_and_trips(
                    wb[legacy_title],
                    service_date=srv_date,
                    default_company_name=(opts.get("company") or None),
                    default_bus_plate=(opts.get("bus") or None),
                    stdout=self.stdout,
                )
            else:
                self.stdout.write(self.style.WARNING(f"Hoja '{legacy_title}' no encontrada (se omite)."))

        self.stdout.write(self.style.SUCCESS("Importación finalizada."))
